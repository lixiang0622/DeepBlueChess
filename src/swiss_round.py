import tkinter as tk
from tkinter import ttk, messagebox, filedialog, simpledialog, Menu
import os
import json
import random
import copy
from datetime import datetime

# 尝试导入 matplotlib
try:
    import matplotlib.pyplot as plt
    from matplotlib.table import Table
    from matplotlib.backends.backend_pdf import PdfPages
    # 适配中文字体
    plt.rcParams['font.sans-serif'] = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS', 'PingFang SC']
    plt.rcParams['axes.unicode_minus'] = False
    HAS_MPL = True
except ImportError:
    HAS_MPL = False

# ==========================================
# 核心逻辑层 (Model)
# ==========================================

class Player:
    def __init__(self, name, student_id="", team="", gender="", category=""):
        self.name = name
        self.student_id = student_id
        self.team = team
        self.gender = gender
        self.category = category
        
        self.number = 0         # 比赛编号
        self.score = 0.0        # 积分
        self.opponents = set()
        self.matches = []
        
        self.wins = 0
        self.draws = 0
        self.losses = 0
        self.sonneborn_berger = 0.0
        self.bye = False
        self.color_history = []
        self.absent_streak = 0
        
        self.disqualified = False # 是否被取消资格/退赛
        self.adjustments = []
        self.rank = 0

    def to_dict(self):
        return {
            "name": self.name, "sid": self.student_id, "team": self.team,
            "gender": self.gender, "cat": self.category, "num": self.number,
            "score": self.score, "opps": list(self.opponents), "matches": self.matches,
            "stats": [self.wins, self.draws, self.losses], "sb": self.sonneborn_berger,
            "bye": self.bye, "colors": self.color_history, "absent": self.absent_streak,
            "dq": self.disqualified, "adj": self.adjustments, "rank": self.rank
        }

    @classmethod
    def from_dict(cls, d):
        p = cls(d["name"], d["sid"], d["team"], d["gender"], d["cat"])
        p.number = d["num"]
        p.score = d["score"]
        p.opponents = set(d["opps"])
        p.matches = d["matches"]
        p.wins, p.draws, p.losses = d["stats"]
        p.sonneborn_berger = d["sb"]
        p.bye = d["bye"]
        p.color_history = d["colors"]
        p.absent_streak = d["absent"]
        p.disqualified = d["dq"]
        p.adjustments = d["adj"]
        p.rank = d["rank"]
        return p

    def add_result(self, opponent_name, score, color, round_no, note=""):
        self.matches.append({
            'round': round_no,
            'opponent': opponent_name,
            'color': color,
            'result': score,
            'note': note
        })
        if opponent_name != "BYE":
            self.opponents.add(opponent_name)
        if color:
            self.color_history.append(color)

        self.score += score
        if score == 1: self.wins += 1
        elif score == 0.5: self.draws += 1
        else: self.losses += 1
            
        if "缺席" in note:
            self.absent_streak += 1
        else:
            self.absent_streak = 0

    def apply_penalty(self, points, reason):
        self.score += points
        self.adjustments.append((points, reason))

    def calculate_sb(self, all_players_dict):
        sb = 0.0
        for m in self.matches:
            opp = m['opponent']
            if opp == "BYE" or opp not in all_players_dict: continue
            opp_obj = all_players_dict[opp]
            if m['result'] == 1: sb += opp_obj.score
            elif m['result'] == 0.5: sb += 0.5 * opp_obj.score
        self.sonneborn_berger = sb

    def get_color_balance(self):
        return self.color_history.count('W') - self.color_history.count('B')

class TournamentEngine:
    def __init__(self):
        self.players = []
        self.rounds_total = 5
        self.current_round = 0
        self.pairings = [] 
        self.results = {} 
        self.history = []
        self.use_teams = False
        self.is_started = False
        self.meta = {"name": "未命名比赛"}
        self.save_dir = ""

    def init_save_dir(self):
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        safe_name = "".join([c for c in self.meta['name'] if c.isalnum() or c in (' ', '_', '-')]).strip()
        folder_name = f"{safe_name}_{ts}"
        self.save_dir = os.path.join(os.getcwd(), folder_name)
        if not os.path.exists(self.save_dir):
            os.makedirs(self.save_dir)
        return self.save_dir

    def save_state(self, stage_suffix):
        if not self.save_dir: self.init_save_dir()
        data = {
            "meta": self.meta,
            "config": {"rounds": self.rounds_total, "use_teams": self.use_teams, "started": self.is_started},
            "status": {"current_round": self.current_round},
            "players": [p.to_dict() for p in self.players],
            "history": self.history,
            "current_pairings": [],
            "current_results": {str(k): v for k, v in self.results.items()}
        }
        p_map = {p.name: i for i, p in enumerate(self.players)}
        cp = []
        for p1, p2_obj, color in self.pairings:
            idx1 = p_map[p1.name]
            idx2 = "BYE" if p2_obj == "BYE" else p_map[p2_obj.name]
            cp.append((idx1, idx2, color))
        data["current_pairings"] = cp

        filename = f"R{self.current_round:02d}_{stage_suffix}.json"
        path = os.path.join(self.save_dir, filename)
        with open(path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        return path

    def load_state(self, filepath):
        with open(filepath, 'r', encoding='utf-8') as f:
            data = json.load(f)
        self.meta = data["meta"]
        self.rounds_total = data["config"]["rounds"]
        self.use_teams = data["config"]["use_teams"]
        self.is_started = data["config"]["started"]
        self.current_round = data["status"]["current_round"]
        self.history = data["history"]
        self.players = [Player.from_dict(d) for d in data["players"]]
        self.pairings = []
        cp = data.get("current_pairings", [])
        for idx1, idx2, color in cp:
            p1 = self.players[idx1]
            p2 = "BYE" if idx2 == "BYE" else self.players[idx2]
            self.pairings.append((p1, p2, color))
        self.results = {}
        cr = data.get("current_results", {})
        for k, v in cr.items():
            self.results[int(k)] = tuple(v)
            
        parent_dir = os.path.dirname(filepath)
        if os.path.exists(parent_dir): self.save_dir = parent_dir
        else: self.init_save_dir()
        return True

    def add_player(self, name, student_id="", team="", gender="", category=""):
        name_check = name
        cnt = 1
        existing = [p.name for p in self.players]
        while name_check in existing:
            cnt += 1
            name_check = f"{name} ({cnt})"
        p = Player(name_check, student_id, team, gender, category)
        p.number = len(self.players) + 1
        self.players.append(p)
        return p

    def remove_player_by_index(self, index):
        """物理删除选手（仅在比赛开始前使用）"""
        if 0 <= index < len(self.players):
            del self.players[index]
            # 重新生成编号
            for i, p in enumerate(self.players): p.number = i + 1
            return True
        return False

    def disqualify_player_by_name(self, name):
        """比赛中途删除（标记为DQ）"""
        for p in self.players:
            if p.name == name:
                p.disqualified = True
                p.adjustments.append((0, "裁判: 手动删除/退赛"))
                return True
        return False

    def rank_players(self):
        pmap = {p.name: p for p in self.players}
        for p in self.players: p.calculate_sb(pmap)
        # 排序规则：积分 > 小分 > 胜局
        self.players.sort(key=lambda x: (x.score, x.sonneborn_berger, x.wins), reverse=True)
        for i, p in enumerate(self.players): p.rank = i + 1
        return self.players

    def generate_pairings(self):
        # 1. 获取有效选手（排除已删除/退赛的）
        active = [p for p in self.players if not p.disqualified]
        
        # 2. 轮空处理
        bye_p = None
        if len(active) % 2 != 0:
            if self.current_round == 0:
                # 第一轮如果要轮空，随机选一个（或按规则选），这里随机
                random.shuffle(active)
                bye_p = active.pop()
            else:
                # 后续轮次，分最低且未轮空过的轮空
                active.sort(key=lambda x: (x.bye, x.score))
                bye_p = active.pop(0)

        # 3. 排序策略
        if self.current_round == 0:
            # === [新功能] 第一轮强制随机 ===
            random.shuffle(active)
        else:
            # === 瑞士轮标准：高分对高分 ===
            active.sort(key=lambda x: x.score, reverse=True)
        
        # 4. 递归回溯配对
        memo = {}
        def solve(rem):
            k = tuple(p.name for p in rem)
            if k in memo: return memo[k]
            if not rem: return []
            
            p1 = rem[0]
            for i in range(1, len(rem)):
                p2 = rem[i]
                if p2.name in p1.opponents: continue
                # 只有非第一轮才严格检查同队，或者可配置。第一轮随机时如果同队碰上也认了？通常还是避开
                if self.use_teams and p1.team and p2.team and p1.team == p2.team: continue
                
                res = solve(rem[1:i] + rem[i+1:])
                if res is not None:
                    # 分配颜色
                    bal1, bal2 = p1.get_color_balance(), p2.get_color_balance()
                    c = 'B' if bal1 > bal2 else ('W' if bal1 < bal2 else ('B' if p1.color_history and p1.color_history[-1]=='W' else 'W'))
                    return [(p1, p2, c)] + res
            memo[k] = None
            return None

        pairs = solve(active)
        
        # 容错：如果严格规则导致无法配对（如同队死锁），尝试放宽限制
        if pairs is None and self.use_teams:
            memo = {}
            def solve_loose(rem):
                k = tuple(p.name for p in rem)
                if k in memo: return memo[k]
                if not rem: return []
                p1 = rem[0]
                for i in range(1, len(rem)):
                    p2 = rem[i]
                    if p2.name in p1.opponents: continue
                    res = solve_loose(rem[1:i] + rem[i+1:])
                    if res is not None: return [(p1, p2, 'W')] + res
                memo[k] = None
                return None
            pairs = solve_loose(active)

        if pairs is None: 
            # 极少数情况，再次打乱重试（针对第一轮死锁）
            if self.current_round == 0:
                random.shuffle(active)
                # 简单两两配对，不再回溯
                pairs = []
                while len(active) >= 2:
                    p1 = active.pop()
                    p2 = active.pop()
                    pairs.append((p1, p2, 'W'))
            else:
                raise Exception("无法生成合法对阵，请尝试人工调整或检查人数。")
        
        self.pairings = [(p1, p2, c) for p1, p2, c in pairs]
        if bye_p:
            self.pairings.append((bye_p, "BYE", None))
            bye_p.bye = True
            
        self.current_round += 1
        self.results = {}
        return self.pairings

    def commit_round(self):
        if len(self.results) < len(self.pairings): return False, "未录入全部结果"
        
        rh = {"round": self.current_round, "matches": []}
        for i, (p1, p2_obj, color) in enumerate(self.pairings):
            s1, s2, note = self.results.get(i, (0,0,""))
            c1 = color
            c2 = 'B' if color=='W' else 'W'
            if not color: c1, c2 = None, None
            
            p2_name = p2_obj.name if isinstance(p2_obj, Player) else "BYE"
            p1.add_result(p2_name, s1, c1, self.current_round, note)
            if isinstance(p2_obj, Player):
                p2_obj.add_result(p1.name, s2, c2, self.current_round, note)
                self._check_dq(p1)
                self._check_dq(p2_obj)
            
            wn = p1.name if c1=='W' else (p2_name if c1=='B' else p1.name)
            bn = p2_name if c1=='W' else (p1.name if c1=='B' else "")
            res_str = f"{s1}-{s2}"
            if note: res_str += f"({note})"
            rh["matches"].append({"white": wn, "black": bn, "result": res_str})
            
        self.history.append(rh)
        return True, "OK"

    def _check_dq(self, p):
        if p.absent_streak >= 2 and not p.disqualified:
            p.disqualified = True
            p.adjustments.append((0, "系统: 连缺自动DQ"))

# ==========================================
# UI 视图层
# ==========================================

class ChessApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("国际象棋赛事管理系统 Pro Max (Random Round 1)")
        self.geometry("1100x850")
        self.engine = TournamentEngine()
        
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Treeview", rowheight=30, font=('Microsoft YaHei', 11))
        style.configure("Treeview.Heading", font=('Microsoft YaHei', 11, 'bold'))
        style.map('Treeview', background=[('selected', '#0078D7')], foreground=[('selected', 'white')])

        self.main_container = ttk.Frame(self)
        self.main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        self.notebook = ttk.Notebook(self.main_container)
        self.notebook.pack(fill=tk.BOTH, expand=True)
        
        self.tab_setup = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_setup, text=" 1. 报名/加载 ")
        self._init_setup_tab()
        
        self.tab_pairing = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_pairing, text=" 2. 比赛进行 ")
        self._init_pairing_tab()
        
        self.tab_standings = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_standings, text=" 3. 排名/裁判 ")
        self._init_standings_tab()

    # --- 辅助：格式化选手显示字符串 ---
    def fmt_player(self, p):
        """返回格式：[#1] 张三 (2.0)"""
        if isinstance(p, Player):
            return f"[#{p.number}] {p.name} ({p.score})"
        return str(p) # BYE

    # --- 通用：导出表格图片 ---
    def export_table_image(self, data, col_labels, title, filename):
        if not HAS_MPL or not self.engine.save_dir: return
        
        fig, ax = plt.subplots(figsize=(12, len(data)*0.4 + 2))
        ax.axis('off')
        
        table = Table(ax, bbox=[0, 0, 1, 1])
        # Header
        for i, col in enumerate(col_labels):
            table.add_cell(0, i, 0.1, 0.1, text=col, loc='center', facecolor='#dddddd')
        # Data
        for row_idx, row in enumerate(data):
            for col_idx, val in enumerate(row):
                table.add_cell(row_idx+1, col_idx, 0.1, 0.1, text=str(val), loc='center')
        
        ax.add_table(table)
        plt.title(title, fontsize=16, pad=20)
        
        save_path = os.path.join(self.engine.save_dir, filename)
        plt.savefig(save_path, bbox_inches='tight', dpi=150)
        plt.close(fig)
        return save_path

    # --- Setup Tab ---
    def _init_setup_tab(self):
        f = self.tab_setup
        
        load_frame = ttk.LabelFrame(f, text="恢复比赛")
        load_frame.pack(fill=tk.X, pady=5)
        ttk.Button(load_frame, text="📂 加载状态文件 (.json)", command=self.action_load_state).pack(pady=10)

        info_f = ttk.LabelFrame(f, text="新比赛设置")
        info_f.pack(fill=tk.X, pady=5)
        ttk.Label(info_f, text="比赛名称:").grid(row=0, column=0, padx=5, pady=10)
        self.ent_name = ttk.Entry(info_f, width=20)
        self.ent_name.grid(row=0, column=1, padx=5)
        ttk.Label(info_f, text="轮数:").grid(row=0, column=2, padx=5)
        self.ent_rounds = ttk.Spinbox(info_f, from_=1, to=20, width=5)
        self.ent_rounds.set(5)
        self.ent_rounds.grid(row=0, column=3, padx=5)
        self.var_use_team = tk.BooleanVar()
        ttk.Checkbutton(info_f, text="启用团体计分", variable=self.var_use_team).grid(row=0, column=4, padx=15)

        input_f = ttk.LabelFrame(f, text="录入选手")
        input_f.pack(fill=tk.X, pady=5)
        ttk.Label(input_f, text="学号:").grid(row=0, column=0)
        self.ent_pid = ttk.Entry(input_f, width=10)
        self.ent_pid.grid(row=0, column=1, padx=5)
        ttk.Label(input_f, text="姓名:").grid(row=0, column=2)
        self.ent_pname = ttk.Entry(input_f, width=10)
        self.ent_pname.grid(row=0, column=3, padx=5)
        ttk.Label(input_f, text="队伍:").grid(row=0, column=4)
        self.ent_pteam = ttk.Entry(input_f, width=10)
        self.ent_pteam.grid(row=0, column=5, padx=5)
        ttk.Label(input_f, text="性别:").grid(row=0, column=6)
        self.cmb_pgender = ttk.Combobox(input_f, values=["男", "女"], width=5)
        self.cmb_pgender.grid(row=0, column=7, padx=5)
        
        bf = ttk.Frame(input_f)
        bf.grid(row=0, column=8, padx=10)
        ttk.Button(bf, text="添加", command=self.action_add_player).pack(side=tk.LEFT, padx=2)
        ttk.Button(bf, text="Excel导入", command=self.action_import_excel).pack(side=tk.LEFT, padx=2)
        ttk.Button(bf, text="删除(报名期)", command=self.action_delete_player).pack(side=tk.LEFT, padx=2)

        lf = ttk.Frame(f)
        lf.pack(fill=tk.BOTH, expand=True)
        cols = ("序号", "学号", "姓名", "队伍", "性别")
        self.tree_players = ttk.Treeview(lf, columns=cols, show='headings')
        for c in cols: self.tree_players.heading(c, text=c)
        self.tree_players.column("序号", width=50, anchor='center')
        self.tree_players.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        
        self.btn_start = ttk.Button(f, text="开始比赛", command=self.action_start_tournament)
        self.btn_start.pack(pady=10, ipady=5)

    # --- Pairing Tab ---
    def _init_pairing_tab(self):
        f = self.tab_pairing
        tb = ttk.Frame(f)
        tb.pack(fill=tk.X, pady=5)
        self.lbl_round = ttk.Label(tb, text="等待开始", font=('Microsoft YaHei', 14, 'bold'))
        self.lbl_round.pack(side=tk.LEFT, padx=10)
        self.btn_next = ttk.Button(tb, text="结算本轮 -> 下一轮", state=tk.DISABLED, command=self.action_next_round)
        self.btn_next.pack(side=tk.RIGHT, padx=10)

        lc = ttk.Frame(f)
        lc.pack(fill=tk.BOTH, expand=True, pady=5)
        cols = ("ID", "台次", "白方", "结果", "黑方", "状态")
        self.tree_pairings = ttk.Treeview(lc, columns=cols, show='headings', selectmode="browse")
        self.tree_pairings.heading("台次", text="台次")
        self.tree_pairings.column("台次", width=60, anchor='center')
        self.tree_pairings.column("ID", width=0, stretch=False)
        self.tree_pairings.heading("白方", text="白方 [编号]姓名 (分)")
        self.tree_pairings.column("白方", width=250, anchor='center')
        self.tree_pairings.heading("结果", text="结果")
        self.tree_pairings.column("结果", width=150, anchor='center')
        self.tree_pairings.heading("黑方", text="黑方 [编号]姓名 (分)")
        self.tree_pairings.column("黑方", width=250, anchor='center')
        self.tree_pairings.heading("状态", text="状态")
        self.tree_pairings.column("状态", width=100, anchor='center')
        
        self.tree_pairings.tag_configure('pending', foreground='black')
        self.tree_pairings.tag_configure('done_normal', foreground='green')
        self.tree_pairings.tag_configure('done_draw', foreground='#00008B')
        self.tree_pairings.tag_configure('done_forfeit', foreground='red')
        self.tree_pairings.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        ip = ttk.LabelFrame(f, text="快速录入")
        ip.pack(fill=tk.X, pady=10, ipady=5)
        btns = [
            ("白胜 1-0", 1, 0, "", "#E8F5E9"), ("黑胜 0-1", 0, 1, "", "#E8F5E9"),
            ("和棋 ½", 0.5, 0.5, "", "#E3F2FD"),
            ("白弃 (0-1F)", 0, 1, "白缺席", "#FFEBEE"), ("黑弃 (1F-0)", 1, 0, "黑缺席", "#FFEBEE"),
            ("双弃 0-0", 0, 0, "双缺席", "#FFCDD2")
        ]
        for txt, s1, s2, n, bg in btns:
            tk.Button(ip, text=txt, font=('Microsoft YaHei', 10), bg=bg, height=2,
                      command=lambda s1=s1, s2=s2, n=n: self.apply_match_result(s1, s2, n)
                     ).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=2)

    # --- Standings Tab ---
    def _init_standings_tab(self):
        f = self.tab_standings
        cols = ("排名", "学号", "姓名", "队伍", "积分", "小分", "胜", "和", "负", "状态")
        self.tree_standings = ttk.Treeview(f, columns=cols, show='headings')
        for c in cols: self.tree_standings.heading(c, text=c); self.tree_standings.column(c, width=80)
        self.tree_standings.pack(fill=tk.BOTH, expand=True, pady=5)
        
        bf = ttk.Frame(f)
        bf.pack(pady=5)
        ttk.Button(bf, text="刷新排名", command=self.refresh_rankings).pack(side=tk.LEFT, padx=5)
        
        menu = Menu(self, tearoff=0)
        menu.add_command(label="违例扣分", command=self.action_referee_penalty)
        menu.add_separator()
        menu.add_command(label="🚫 彻底删除/退赛 (Disqualify)", command=lambda: self.action_referee_dq(True))
        menu.add_command(label="✅ 恢复资格 (Re-qualify)", command=lambda: self.action_referee_dq(False))
        self.tree_standings.bind("<Button-3>", lambda e: self.do_popup(e, menu))

    def do_popup(self, event, menu):
        row = self.tree_standings.identify_row(event.y)
        if row:
            self.tree_standings.selection_set(row)
            menu.tk_popup(event.x_root, event.y_root)

    # ================= 业务动作 =================

    def action_load_state(self):
        f = filedialog.askopenfilename(filetypes=[("JSON", "*.json")])
        if not f: return
        try:
            if self.engine.load_state(f):
                # 恢复UI
                self.ent_name.delete(0, tk.END); self.ent_name.insert(0, self.engine.meta['name'])
                self.ent_rounds.set(self.engine.rounds_total)
                self.var_use_team.set(self.engine.use_teams)
                self.btn_start.config(state=tk.DISABLED)
                self._refresh_player_list()
                
                if self.engine.pairings:
                    self.lbl_round.config(text=f"第 {self.engine.current_round} / {self.engine.rounds_total} 轮 (恢复)")
                    self._update_pairings_ui(restore=True)
                    self.notebook.select(self.tab_pairing)
                    self.btn_next.config(state=tk.NORMAL)
                else:
                    self.refresh_rankings()
                    self.notebook.select(self.tab_standings)
                
                messagebox.showinfo("成功", f"已恢复到第 {self.engine.current_round} 轮")
        except Exception as e:
            messagebox.showerror("加载失败", str(e))

    def action_add_player(self):
        if self.engine.is_started: return
        name = self.ent_pname.get().strip()
        if name:
            self.engine.add_player(name, self.ent_pid.get(), self.ent_pteam.get(), self.cmb_pgender.get())
            self._refresh_player_list()
            self.ent_pname.delete(0, tk.END); self.ent_pid.delete(0, tk.END)

    def action_delete_player(self):
        # 这是Tab1的删除按钮，仅在比赛前有效
        if self.engine.is_started: 
            messagebox.showinfo("提示", "比赛已开始，请前往'排名/裁判'页面右键选择'删除/退赛'。")
            return
        sel = self.tree_players.selection()
        for i in sorted([self.tree_players.index(s) for s in sel], reverse=True):
            self.engine.remove_player_by_index(i)
        self._refresh_player_list()

    def _refresh_player_list(self):
        for i in self.tree_players.get_children(): self.tree_players.delete(i)
        for p in self.engine.players:
            self.tree_players.insert("", tk.END, values=(p.number, p.student_id, p.name, p.team, p.gender))

    def action_import_excel(self):
        if self.engine.is_started: return
        f = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx")])
        if not f: return
        try:
            import openpyxl
            wb = openpyxl.load_workbook(f)
            ws = wb.active
            headers = {}
            for r in range(1, 6):
                vals = [str(c.value).strip() if c.value else "" for c in ws[r]]
                if "姓名" in vals or "Name" in vals:
                    for i, v in enumerate(vals, 1):
                        if v in ["姓名","Name"]: headers['name']=i
                        elif v in ["学号","ID"]: headers['id']=i
                        elif v in ["队伍","Team"]: headers['team']=i
                        elif v in ["性别","Gender"]: headers['gender']=i
                    break
            if 'name' not in headers: return
            for r in range(2, ws.max_row+1):
                n = ws.cell(r, headers['name']).value
                if n:
                    sid = ws.cell(r, headers['id']).value if 'id' in headers else ""
                    tm = ws.cell(r, headers['team']).value if 'team' in headers else ""
                    gd = ws.cell(r, headers['gender']).value if 'gender' in headers else ""
                    self.engine.add_player(str(n), str(sid or ""), str(tm or ""), str(gd or ""))
            self._refresh_player_list()
        except: pass

    def action_start_tournament(self):
        if len(self.engine.players) < 2: return
        self.engine.is_started = True
        self.engine.rounds_total = int(self.ent_rounds.get())
        self.engine.use_teams = self.var_use_team.get()
        self.engine.meta['name'] = self.ent_name.get()
        self.engine.init_save_dir()
        
        self.btn_start.config(state=tk.DISABLED)
        self.notebook.select(self.tab_pairing)
        self.run_new_round()

    def run_new_round(self):
        try:
            pairings = self.engine.generate_pairings()
            r = self.engine.current_round
            self.lbl_round.config(text=f"第 {r} / {self.engine.rounds_total} 轮")
            
            # UI Update
            self._update_pairings_ui()
            self.btn_next.config(state=tk.NORMAL)
            
            # Auto Save Image
            data_img = []
            for i, (p1, p2_obj, color) in enumerate(pairings):
                # 导出图片时带编号
                p2_name = "BYE (轮空)" if p2_obj == "BYE" else self.fmt_player(p2_obj)
                p1_name = self.fmt_player(p1)
                
                w, b = (p1_name, p2_name) if color == 'W' else (p2_name, p1_name)
                if p2_obj == "BYE": w, b = p1_name, "BYE"
                data_img.append([i+1, w, "vs", b])
            
            self.export_table_image(data_img, ["Table", "White", "", "Black"], 
                                   f"Round {r} Pairings", f"R{r:02d}_Pairings.png")
            self.engine.save_state("Start")
            
        except Exception as e:
            messagebox.showerror("错误", str(e))

    def _update_pairings_ui(self, restore=False):
        for i in self.tree_pairings.get_children(): self.tree_pairings.delete(i)
        
        for i, (p1, p2_obj, color) in enumerate(self.engine.pairings):
            if p2_obj == "BYE":
                if not restore: self.engine.results[i] = (1, 0, "BYE")
                vals = (i, i+1, self.fmt_player(p1), "1-0 (BYE)", "轮空", "系统录入")
                self.tree_pairings.insert("", tk.END, values=vals, tags=('done_normal',))
            else:
                w = self.fmt_player(p1) if color == 'W' else self.fmt_player(p2_obj)
                b = self.fmt_player(p2_obj) if color == 'W' else self.fmt_player(p1)
                
                res = self.engine.results.get(i)
                if res:
                    s1, s2, note = res
                    if color == 'W': ls, rs = s1, s2
                    else: ls, rs = s2, s1
                    
                    res_txt = f"{ls}-{rs}"
                    if note: res_txt += " F"
                    self.tree_pairings.insert("", tk.END, values=(i, i+1, w, res_txt, b, "已恢复"), tags=('done_normal',))
                else:
                    self.tree_pairings.insert("", tk.END, values=(i, i+1, w, "", b, "待录入"), tags=('pending',))

    def apply_match_result(self, s1, s2, note):
        sel = self.tree_pairings.selection()
        if not sel: return
        iid = sel[0]
        pidx = int(self.tree_pairings.item(iid, "values")[0])
        p1, p2, color = self.engine.pairings[pidx]
        if p2 == "BYE": return
        
        if color == 'W': final_s1, final_s2 = s1, s2
        else: final_s1, final_s2 = s2, s1
            
        self.engine.results[pidx] = (final_s1, final_s2, note)
        
        res_str = f"{s1}-{s2}"
        if note: res_str += " (F)"
        tag = 'done_forfeit' if note else ('done_draw' if s1==s2 else 'done_normal')
        
        vals = list(self.tree_pairings.item(iid, "values"))
        vals[3] = res_str
        vals[5] = "已录入"
        self.tree_pairings.item(iid, values=vals, tags=(tag,))
        
        next_id = self.tree_pairings.next(iid)
        if next_id: 
            self.tree_pairings.selection_set(next_id)
            self.tree_pairings.see(next_id)

    def action_next_round(self):
        ok, msg = self.engine.commit_round()
        if not ok: return messagebox.showwarning("提示", msg)
        
        r = self.engine.current_round
        
        # 保存结果图片
        res_data = []
        for i, (p1, p2, c) in enumerate(self.engine.pairings):
            s1, s2, n = self.engine.results[i]
            p1n = self.fmt_player(p1)
            p2n = "BYE" if p2=="BYE" else self.fmt_player(p2)
            
            if c == 'W': w, b, res = p1n, p2n, f"{s1}-{s2}"
            else: w, b, res = p2n, p1n, f"{s2}-{s1}"
            
            if p2=="BYE": w, b, res = p1n, "BYE", "1-0"
            if n: res += " (F)"
            res_data.append([i+1, w, res, b])
            
        self.export_table_image(res_data, ["Table", "White", "Result", "Black"], 
                                f"Round {r} Results", f"R{r:02d}_Results.png")
        
        self.refresh_rankings()
        ranked = self.engine.players
        rank_data = []
        for p in ranked:
            st = "DQ" if p.disqualified else ("Ab" if p.absent_streak else "")
            rank_data.append([p.rank, p.number, p.name, p.team, p.score, f"{p.sonneborn_berger:.2f}", st])
        
        self.export_table_image(rank_data, ["Rank", "No.", "Name", "Team", "Pts", "SB", "Status"], 
                                f"Standings After Round {r}", f"R{r:02d}_Rankings.png")
        
        self.engine.save_state("End")
        
        if self.engine.current_round >= self.engine.rounds_total:
            messagebox.showinfo("完赛", "比赛结束！正在生成汇总PDF...")
            self.generate_final_pdf()
            self.notebook.select(self.tab_standings)
            self.btn_next.config(state=tk.DISABLED)
        else:
            if messagebox.askyesno("提示", "本轮结束。是否开始下一轮？"):
                self.run_new_round()

    def refresh_rankings(self):
        ranked = self.engine.rank_players()
        for i in self.tree_standings.get_children(): self.tree_standings.delete(i)
        for p in ranked:
            # 状态显示调整
            st = "DQ (退赛)" if p.disqualified else ("缺席" if p.absent_streak else "正常")
            vals = (p.rank, p.student_id, p.name, p.team, p.score, f"{p.sonneborn_berger:.2f}", p.wins, p.draws, p.losses, st)
            self.tree_standings.insert("", tk.END, values=vals)

    def action_referee_penalty(self):
        sel = self.tree_standings.selection()
        if not sel: return
        name = self.tree_standings.item(sel[0], "values")[2]
        p = next((x for x in self.engine.players if x.name == name), None)
        if p:
            v = simpledialog.askfloat("扣分", "分数:")
            if v: p.apply_penalty(v, "裁判")
            self.refresh_rankings()

    def action_referee_dq(self, val):
        sel = self.tree_standings.selection()
        if not sel: return
        name = self.tree_standings.item(sel[0], "values")[2]
        
        # [新功能] 比赛中途删除/DQ
        if val: # 如果是DQ
            if messagebox.askyesno("确认", f"确定要让 {name} 退赛/删除吗？\n他将保留历史成绩，但不再参与后续排位。"):
                self.engine.disqualify_player_by_name(name)
        else: # 恢复
            p = next((x for x in self.engine.players if x.name == name), None)
            if p: p.disqualified = False
            
        self.refresh_rankings()

    def generate_final_pdf(self):
        if not HAS_MPL or not self.engine.save_dir: return
        
        pdf_path = os.path.join(self.engine.save_dir, "Final_Report_Full.pdf")
        player_map = {p.name: p for p in self.engine.players}
        
        with PdfPages(pdf_path) as pdf:
            # Page 1: Final Standings
            fig, ax = plt.subplots(figsize=(11.69, 8.27))
            ax.axis('off')
            plt.title(f"Final Standings - {self.engine.meta['name']}", fontsize=20, pad=20)
            
            cols = ["Rank", "No.", "Name", "Team", "Points", "SB", "W-D-L"]
            data = []
            for p in self.engine.players:
                data.append([p.rank, p.number, p.name, p.team, p.score, f"{p.sonneborn_berger:.2f}", f"{p.wins}-{p.draws}-{p.losses}"])
            
            table = Table(ax, bbox=[0.05, 0.1, 0.9, 0.8])
            for i, c in enumerate(cols): table.add_cell(0, i, 0.1, 0.05, text=c, loc='center', facecolor='#cccccc')
            for r, row in enumerate(data):
                for c, val in enumerate(row):
                    table.add_cell(r+1, c, 0.1, 0.05, text=str(val), loc='center')
            ax.add_table(table)
            pdf.savefig(fig)
            plt.close(fig)
            
            # Page 2+: Player Cards
            players = sorted(self.engine.players, key=lambda x: x.rank)
            for p in players:
                fig, ax = plt.subplots(figsize=(11.69, 8.27))
                ax.axis('off')
                
                header_text = f"Player Card: #{p.number} {p.name} (Rank: {p.rank}, Pts: {p.score})"
                plt.text(0.05, 0.95, header_text, fontsize=16, weight='bold')
                
                p_cols = ["Round", "Color", "Opponent", "Opp.#", "Opp.Rank", "Opp.Pts", "Result", "Note"]
                p_data = []
                
                sorted_matches = sorted(p.matches, key=lambda x: x['round'])
                for m in sorted_matches:
                    c = m['color'] if m['color'] else "-"
                    res = m['result']
                    opp_name = m['opponent']
                    note = m['note']
                    
                    if opp_name == "BYE":
                        o_num, o_rank, o_pts = "-", "-", "-"
                    else:
                        opp_obj = player_map.get(opp_name)
                        if opp_obj:
                            o_num = opp_obj.number
                            o_rank = opp_obj.rank
                            o_pts = opp_obj.score
                        else:
                            o_num, o_rank, o_pts = "?", "?", "?"
                            
                    p_data.append([m['round'], c, opp_name, o_num, o_rank, o_pts, res, note])
                
                sub_table = Table(ax, bbox=[0.05, 0.5, 0.9, 0.4])
                widths = [0.08, 0.08, 0.20, 0.08, 0.1, 0.1, 0.1, 0.15]
                for i, col_name in enumerate(p_cols):
                    sub_table.add_cell(0, i, widths[i], 0.1, text=col_name, loc='center', facecolor='#eeeeee')
                for r, row in enumerate(p_data):
                    for c_idx, val in enumerate(row):
                        sub_table.add_cell(r+1, c_idx, widths[c_idx], 0.1, text=str(val), loc='center')
                        
                ax.add_table(sub_table)
                pdf.savefig(fig)
                plt.close(fig)
                
        messagebox.showinfo("PDF生成", f"已生成详细汇总报告：\n{pdf_path}")
        os.startfile(self.engine.save_dir)

if __name__ == "__main__":
    app = ChessApp()
    app.mainloop()